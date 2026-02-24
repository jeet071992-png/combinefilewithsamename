"""
Core module for combinefilewithsamename
Handles Excel workbook combining with popup file selector
"""

import os
import shutil
import tkinter as tk
from tkinter import filedialog, messagebox

# ─────────────────────────────────────────
# VBA Code (embedded as string)
# ─────────────────────────────────────────
VBA_CODE = '''
Sub CombineWorkbooks()

    Dim dlgOpen     As FileDialog
    Dim srcWB       As Workbook
    Dim destWB      As Workbook
    Dim srcSheet    As Worksheet
    Dim selectedFile As Variant
    Dim fileCount   As Integer
    Dim sheetCount  As Integer
    Dim newName     As String

    Set destWB = ThisWorkbook

    Set dlgOpen = Application.FileDialog(msoFileDialogFilePicker)

    With dlgOpen
        .Title = "Select Workbooks to Combine"
        .AllowMultiSelect = True
        .Filters.Clear
        .Filters.Add "All Excel Formats", "*.xlsx; *.xls; *.xlsm; *.xlsb; *.csv"
        .Filters.Add "Excel Files", "*.xlsx; *.xls; *.xlsm; *.xlsb"
        .Filters.Add "CSV Files", "*.csv"
        .FilterIndex = 1

        If .Show = False Then
            MsgBox "No files selected. Operation cancelled.", vbInformation, "Cancelled"
            Exit Sub
        End If

        fileCount = 0
        sheetCount = 0
        Application.ScreenUpdating = False

        For Each selectedFile In .SelectedItems
            fileCount = fileCount + 1
            Set srcWB = Workbooks.Open(selectedFile, ReadOnly:=True)

            For Each srcSheet In srcWB.Sheets
                srcSheet.Copy After:=destWB.Sheets(destWB.Sheets.Count)
                newName = srcSheet.Name & "_" & fileCount
                On Error Resume Next
                destWB.Sheets(destWB.Sheets.Count).Name = newName
                On Error GoTo 0
                sheetCount = sheetCount + 1
            Next srcSheet

            srcWB.Close SaveChanges:=False
        Next selectedFile
    End With

    Application.ScreenUpdating = True

    MsgBox "Successfully Combined!" & vbNewLine & vbNewLine & _
           "Files Combined : " & fileCount & vbNewLine & _
           "Sheets Added   : " & sheetCount, _
           vbInformation, "Combine Complete"

End Sub
'''


def combine_excel_files(output_path=None):
    """
    Open popup to select Excel files and combine them using openpyxl.
    Pure Python approach (no Excel needed).
    
    Args:
        output_path (str): Where to save combined file. If None, asks user.
    """
    try:
        import openpyxl
    except ImportError:
        print("Installing openpyxl...")
        os.system("pip install openpyxl")
        import openpyxl

    # Hide tkinter root window
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)

    # File selection popup
    files = filedialog.askopenfilenames(
        title="Select Excel Files to Combine",
        filetypes=[
            ("All Excel Formats", "*.xlsx *.xls *.xlsm *.xlsb"),
            ("Excel Files", "*.xlsx"),
            ("Legacy Excel", "*.xls"),
            ("Macro Excel", "*.xlsm"),
            ("All Files", "*.*"),
        ]
    )

    if not files:
        messagebox.showinfo("Cancelled", "No files selected. Operation cancelled.")
        root.destroy()
        return

    # Output path
    if output_path is None:
        output_path = filedialog.asksaveasfilename(
            title="Save Combined File As",
            defaultextension=".xlsx",
            filetypes=[("Excel File", "*.xlsx")]
        )

    if not output_path:
        messagebox.showinfo("Cancelled", "Save location not selected.")
        root.destroy()
        return

    # Combine workbooks
    combined_wb = openpyxl.Workbook()
    combined_wb.remove(combined_wb.active)  # Remove default empty sheet

    sheet_count = 0
    file_count = 0

    for file_path in files:
        file_count += 1
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            file_name = os.path.splitext(os.path.basename(file_path))[0]

            for sheet_name in wb.sheetnames:
                ws_src = wb[sheet_name]
                new_name = f"{sheet_name}_{file_count}"

                # Handle duplicate names
                counter = 1
                base_name = new_name
                while new_name in combined_wb.sheetnames:
                    new_name = f"{base_name}_{counter}"
                    counter += 1

                ws_dest = combined_wb.create_sheet(title=new_name)

                for row in ws_src.iter_rows():
                    for cell in row:
                        ws_dest[cell.coordinate].value = cell.value

                sheet_count += 1

        except Exception as e:
            messagebox.showwarning("Warning", f"Could not read file:\n{file_path}\n\nError: {str(e)}")

    combined_wb.save(output_path)

    messagebox.showinfo(
        "Success ✅",
        f"Successfully Combined!\n\n"
        f"Files Combined : {file_count}\n"
        f"Sheets Added   : {sheet_count}\n\n"
        f"Saved at:\n{output_path}"
    )

    root.destroy()
    print(f"\n✅ Done! Combined file saved at: {output_path}")


def inject_vba(excel_path=None):
    """
    Copy VBA code to clipboard so user can paste in Excel VBA editor.
    
    Args:
        excel_path: Not used currently. Future: auto-inject via win32com.
    """
    try:
        import pyperclip
        pyperclip.copy(VBA_CODE)
        print("\n✅ VBA Code copied to clipboard!")
        print("Now:")
        print("  1. Open Excel")
        print("  2. Press Alt + F11")
        print("  3. Insert > Module")
        print("  4. Paste (Ctrl+V)")
        print("  5. Press F5 to Run!\n")
    except ImportError:
        print("\n📋 VBA Code to paste in Excel:\n")
        print(VBA_CODE)
        print("\nSteps:")
        print("  1. Open Excel → Alt+F11")
        print("  2. Insert > Module → Paste above code")
        print("  3. Press F5 to Run!")


def run():
    """
    Main entry point. Shows menu to choose action.
    """
    print("=" * 50)
    print("  combinefilewithsamename")
    print("  Excel Workbook Combiner Tool")
    print("=" * 50)
    print("\nWhat do you want to do?")
    print("  1. Combine Excel files (Python - no Excel needed)")
    print("  2. Get VBA code for Excel macro")
    print("  3. Exit")
    
    choice = input("\nEnter choice (1/2/3): ").strip()
    
    if choice == "1":
        combine_excel_files()
    elif choice == "2":
        inject_vba()
    elif choice == "3":
        print("Bye!")
    else:
        print("Invalid choice. Running Excel combiner...")
        combine_excel_files()
