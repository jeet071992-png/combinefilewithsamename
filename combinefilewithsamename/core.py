import os
import tkinter as tk
from tkinter import filedialog, messagebox

VBA_CODE = """
Sub CombineWorkbooks()
    Dim dlgOpen As FileDialog
    Dim srcWB As Workbook
    Dim destWB As Workbook
    Dim srcSheet As Worksheet
    Dim selectedFile As Variant
    Dim fileCount As Integer
    Dim sheetCount As Integer
    Dim newName As String
    Set destWB = ThisWorkbook
    Set dlgOpen = Application.FileDialog(msoFileDialogFilePicker)
    With dlgOpen
        .Title = "Select Workbooks to Combine"
        .AllowMultiSelect = True
        .Filters.Clear
        .Filters.Add "All Excel Formats", "*.xlsx; *.xls; *.xlsm; *.xlsb; *.csv"
        .FilterIndex = 1
        If .Show = False Then
            MsgBox "No files selected.", vbInformation, "Cancelled"
            Exit Sub
        End If
        fileCount = 0
        sheetCount = 0
        Application.ScreenUpdating = False
        For Each selectedFile In .SelectedItems
            fileCount = fileCount + 1
            Set srcWB = Workbooks.Open(selectedFile, ReadOnly:=True)
            For Each srcSheet In srcWB.Sheets
                srcSheet.Copy After:=destWB.Sheets(destWB.Sheets.Count)
                newName = srcSheet.Name & "_" & fileCount
                On Error Resume Next
                destWB.Sheets(destWB.Sheets.Count).Name = newName
                On Error GoTo 0
                sheetCount = sheetCount + 1
            Next srcSheet
            srcWB.Close SaveChanges:=False
        Next selectedFile
    End With
    Application.ScreenUpdating = True
    MsgBox "Combined!" & vbNewLine & "Files: " & fileCount & vbNewLine & "Sheets: " & sheetCount, vbInformation, "Done"
End Sub
"""

def combine_excel_files(output_path=None):
    try:
        import openpyxl
    except ImportError:
        os.system("pip install openpyxl")
        import openpyxl

    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)

    files = filedialog.askopenfilenames(
        title="Select Excel Files to Combine",
        filetypes=[("All Excel Formats", "*.xlsx *.xls *.xlsm *.xlsb"), ("All Files", "*.*")]
    )

    if not files:
        messagebox.showinfo("Cancelled", "No files selected.")
        root.destroy()
        return

    if output_path is None:
        output_path = filedialog.asksaveasfilename(
            title="Save Combined File As",
            defaultextension=".xlsx",
            filetypes=[("Excel File", "*.xlsx")]
        )

    if not output_path:
        root.destroy()
        return

    combined_wb = openpyxl.Workbook()
    combined_wb.remove(combined_wb.active)

    sheet_count = 0
    file_count = 0

    for file_path in files:
        file_count += 1
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                ws_src = wb[sheet_name]
                new_name = f"{sheet_name}_{file_count}"
                counter = 1
                base_name = new_name
                while new_name in combined_wb.sheetnames:
                    new_name = f"{base_name}_{counter}"
                    counter += 1
                ws_dest = combined_wb.create_sheet(title=new_name)
                for row in ws_src.iter_rows():
                    for cell in row:
                        ws_dest[cell.coordinate].value = cell.value
                sheet_count += 1
        except Exception as e:
            messagebox.showwarning("Warning", f"Could not read:\n{file_path}\n\n{str(e)}")

    combined_wb.save(output_path)
    messagebox.showinfo("Success!", f"Files: {file_count}\nSheets: {sheet_count}\nSaved at:\n{output_path}")
    root.destroy()
    print(f"Done! Saved at: {output_path}")

def inject_vba():
    try:
        import pyperclip
        pyperclip.copy(VBA_CODE)
        print("VBA Code copied to clipboard!")
    except ImportError:
        print(VBA_CODE)
    print("\nSteps: Open Excel > Alt+F11 > Insert > Module > Paste > F5")

def run():
    print("=" * 40)
    print("  combinefilewithsamename")
    print("=" * 40)
    print("1. Combine Excel files")
    print("2. Get VBA code")
    print("3. Exit")
    choice = input("\nChoice (1/2/3): ").strip()
    if choice == "1":
        combine_excel_files()
    elif choice == "2":
        inject_vba()
    else:
        print("Bye!")
